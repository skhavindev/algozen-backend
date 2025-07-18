import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os
import datetime
import win32com.client
import pythoncom
from dateutil.parser import parse
import uuid
import tkinter as tk
from tkinter import simpledialog

pythoncom.CoInitialize()

DESKTOP_PATH = r"C:\Users\USER\Desktop"
BILLING_BASE_PATH = os.path.join(DESKTOP_PATH, "Billing")
MAIL_IDS_PATH = os.path.join(DESKTOP_PATH, "Content.xlsx")


def convert_xls_to_xlsx(input_path, output_path):
    """Convert an .xls file to .xlsx format using Excel COM automation."""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(input_path)
        wb.SaveAs(output_path, FileFormat=51)  # 51 = .xlsx
        wb.Close()
        excel.Quit()
        print(f"Converted {input_path} to {output_path}")
        return True
    except Exception as e:
        print(f"Error converting {input_path} to .xlsx: {e}")
        return False


def get_due_date():
    """Prompt user for due date between today and today + 4 days."""
    root = tk.Tk()
    root.withdraw()
    today = datetime.date.today()
    max_date = today + datetime.timedelta(days=4)
    while True:
        due_date_str = simpledialog.askstring(
            "Input",
            f"Enter due date (MM/DD/YYYY, between "
            f"{today.strftime('%m/%d/%Y')} and {max_date.strftime('%m/%d/%Y')}):"
        )
        try:
            due_date = datetime.datetime.strptime(due_date_str, "%m/%d/%Y").date()
            if today <= due_date <= max_date:
                return due_date.strftime("%d/%m/%Y")
            else:
                print(
                    f"Date must be between {today.strftime('%m/%d/%Y')} "
                    f"and {max_date.strftime('%m/%d/%Y')}."
                )
        except ValueError:
            print("Invalid date format. Use MM/DD/YYYY.")


def create_pivot_table(df, sheet, party_col, activity_col, amount_col):
    """Write a pivot table to the worksheet."""
    pivot_data = df.pivot_table(
        values=amount_col,
        index=party_col,
        columns=activity_col,
        aggfunc="sum",
        fill_value=0
    )

    # Column headers
    for col_num, value in enumerate(pivot_data.columns, start=1):
        sheet.cell(row=5, column=col_num + 1).value = value

    # Index (row headers) and data
    for row_num, (index, row) in enumerate(pivot_data.iterrows(), start=6):
        sheet.cell(row=row_num, column=1).value = index
        for col_num, value in enumerate(row, start=1):
            sheet.cell(row=row_num, column=col_num + 1).value = value

    # Excel table
    table_ref = f"A5:{get_column_letter(len(pivot_data.columns)+1)}{len(pivot_data)+5}"
    table = openpyxl.worksheet.table.Table(displayName="PivotTable", ref=table_ref)
    table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    sheet.add_table(table)
    return pivot_data


def save_detail_report(df, party, activity, folder, div):
    """Save detailed report for a party and activity."""
    filtered_df = df[(df["Party Name"] == party) & (df["Activity"] == activity)]
    if filtered_df.empty:
        return None

    output_file = os.path.join(folder, f"{activity}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = activity

    # --------------------------------------------------------------
    # HEADERS
    # --------------------------------------------------------------
    headers = filtered_df.columns.tolist()
    if activity != "BONDED AREA DEMURRAGE":
        headers = [h for h in headers if h not in
                   ["Preodic Bill No", "Master Bill No", "Dem Days"]]

    for col, header in enumerate(headers, 1):
        ws[f"{get_column_letter(col)}1"] = header

    # --------------------------------------------------------------
    # DATA
    # --------------------------------------------------------------
    for row, data in enumerate(filtered_df.values, 2):
        for col, value in enumerate(data, 1):
            if col <= len(headers):
                ws[f"{get_column_letter(col)}{row}"] = value

    # --------------------------------------------------------------
    #  NEW: put Excel formula for Dem Days on BONDED AREA DEMURRAGE
    # --------------------------------------------------------------
    if activity == "BONDED AREA DEMURRAGE":
        try:
            dem_col = headers.index("Dem Days") + 1
            gp_col = headers.index("GP/ATD Dt") + 1
            seg_col = headers.index("Seg/Bag Dt") + 1

            for r in range(2, len(filtered_df) + 2):
                dem_cell = f"{get_column_letter(dem_col)}{r}"
                gp_cell = f"{get_column_letter(gp_col)}{r}"
                seg_cell = f"{get_column_letter(seg_col)}{r}"
                # =ROUNDUP(MAX(0, GP - SEG - 1.5),0)
                ws[dem_cell] = (
                    f"=ROUNDUP(MAX(0,{gp_cell}-{seg_cell}-1.5),0)"
                )
        except ValueError:
            # Required columns not found – skip silently
            pass
    # --------------------------------------------------------------

    # Autofit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Excel table
    table_ref = f"A1:{get_column_letter(len(headers))}{len(filtered_df)+1}"
    table = openpyxl.worksheet.table.Table(
        displayName=f"Table_{activity.replace(' ', '_')}",
        ref=table_ref
    )
    table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)

    wb.save(output_file)
    return output_file


def send_email(party, email, period, folder, due_date, amounts):
    """Send email with attachments using Outlook."""
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = email
    mail.CC = "psharsha@AAI.AERO"
    mail.Subject = f"Performa Bill Details-EXPORT Period-{period} - {party}"

    # HTML body
    html_body = (
        f"<html><body>"
        f"<p>Dear Sir/Madam,</p>"
        f"<p>Please find attached the Performa Bill Details for the period {period}.</p>"
        f"<table border='1'><tr><th>Activity</th><th>Amount (Rs.)</th></tr>"
    )
    for act, amt in amounts.items():
        if amt > 0:
            html_body += f"<tr><td>{act}</td><td>{amt:,.0f}</td></tr>"
    html_body += (
        "</table>"
        f"<p>Kindly verify the details by {due_date}. "
        "For queries, contact us at receivablesmaa@aaiclas.aero or 044-22560551.</p>"
        "<p>Best regards,<br>AAICLAS</p>"
        "</body></html>"
    )
    mail.HTMLBody = html_body

    # Attachments
    for file in os.listdir(folder):
        if file.endswith(".xlsx"):
            mail.Attachments.Add(os.path.join(folder, file))

    mail.Send()


def main():
    file_path = (
        r"C:\Users\User\Downloads\EXPORT_INVOICE_SUMMARY-'01062025_15062025' (9).xls"
    )
    xlsx_path = file_path.replace(".xls", ".xlsx")

    if not os.path.exists(file_path):
        print(f"Error: File {file_path} does not exist.")
        return
    print(f"Attempting to read file: {file_path}")

    if not convert_xls_to_xlsx(file_path, xlsx_path):
        print("Failed to convert file to .xlsx.")
        return

    try:
        df = pd.read_excel(xlsx_path, engine="openpyxl")
        print("File read successfully.")
    except Exception as e:
        print(f"Error reading {xlsx_path}: {e}")
        return

    # Division & period
    file_name = os.path.basename(file_path)
    div = file_name[:6]  # EXPORT
    period_raw = file_name[24:41]  # 01062025_15062025
    period = (
        f"{period_raw[:2]}-{period_raw[2:4]}-{period_raw[4:8]}"
        f" to {period_raw[9:11]}-{period_raw[11:13]}-{period_raw[13:]}"
    )

    # Billing folder
    billing_folder = os.path.join(BILLING_BASE_PATH, period, div)
    os.makedirs(billing_folder, exist_ok=True)

    # Clean / compute Dem Days
    df = df.drop(columns=["Bill No", "Bill Date", "Invoice No", "Comm Type"])
    if div == "EXPORT":

        def calc_dem(row):
            if pd.isna(row["GP/ATD Dt"]) or pd.isna(row["Seg/Bag Dt"]):
                return 0
            return max(
                0, (row["GP/ATD Dt"] - row["Seg/Bag Dt"]).days - 1.5
            )

        df["Dem Days"] = df.apply(calc_dem, axis=1)
    df = df.drop(columns=["Preodic Bill No"])

    # Pivot
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pivot"
    pivot = create_pivot_table(df, ws, "Party Name", "Activity", "Amount")
    wb.save(os.path.join(billing_folder, "PivotTable.xlsx"))

    # Emails
    mail_df = pd.read_excel(MAIL_IDS_PATH, sheet_name="Sheet2")
    mail_dict = dict(zip(mail_df.iloc[:, 0], mail_df.iloc[:, 1]))

    for party in pivot.index:
        if party not in mail_dict or pd.isna(mail_dict[party]):
            print(f"Error: Email not found for {party}")
            return

    due_date = get_due_date()

    for party in pivot.index:
        party_folder = os.path.join(billing_folder, party)
        os.makedirs(party_folder, exist_ok=True)

        amounts = pivot.loc[party].to_dict()

        for activity in pivot.columns:
            if amounts.get(activity, 0) > 0:
                save_detail_report(df, party, activity, party_folder, div)

        send_email(
            party, mail_dict[party], period, party_folder, due_date, amounts
        )

    print("Processing complete.")


if __name__ == "__main__":
    main()

